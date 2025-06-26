import openpyxl as xl
from openpyxl.styles import Border, Side, Font, Alignment
import openpyxl.worksheet.worksheet as wsxl
import pandas as pd
from PIL import Image as PILImage
from openpyxl.drawing.image import Image
import io

from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText
from plotly.graph_objs import Figure

from src.constants import TEMP_PATH, VARNAMES_ANALYSIS

import pdb

EXCEL_CELL_LENGTH_COEF = 1.5


def copy_sheet(source_wb: xl.Workbook, source_sheet_name: str, target_wb: xl.Workbook, target_sheet_name: str):
    source_sheet = source_wb[source_sheet_name]
    target_sheet = target_wb.create_sheet(title=target_sheet_name)

    for row in source_sheet.iter_rows(values_only=False):
        for cell in row:
            new_cell = target_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = xl.styles.Font(**cell.font.__dict__)
                new_cell.border = xl.styles.Border(**cell.border.__dict__)
                new_cell.fill = xl.styles.PatternFill(**cell.fill.__dict__)
                new_cell.protection = xl.styles.Protection(**cell.protection.__dict__)
                new_cell.alignment = xl.styles.Alignment(**cell.alignment.__dict__)

    for merged_cell in source_sheet.merged_cells:
        target_sheet.merge_cells(range_string=merged_cell.coord)


'''
---------------------------------------------------------------------------------------------
Data dict: 
data = {
    'fields': {
        "placement1": {
            "stat_params": {
                "var": {
                    "distribution": str,
                    "params": str,
                },    
            },
            "init_data": dict,
            "prod_profile_init_data": dict,
            "risks_params": dict,
            "risks_kriterias": {
                "seismic...": {
                    "kriteria":,
                    "value":,
                    "weight":,
                },
                ...
            },
            "study_coef": float,
            "profiles_report": pd.DataFrame,
            "images": {"hist": bytes, "tornado": bytes, "profile": bytes},
        }, 
        "placement2": {
            ...
        },
        "placement3": {
            ...
        },
        ...
    },
    'comparison': {
        'comparison_values': {},
        'comparison_images': {},
---------------------------------------------------------------------------------------------
'''
def insert_to_sheet(field_name: str, data: dict, sheet: wsxl.Worksheet):
    vars = ['area', 'effective_thickness', 'porosity_coef', 'gas_saturation_coef']
    stat_params: dict = data.get('stat_params', dict())
    for index in range(0, 4, 1):
        param = stat_params.get(vars[index], dict())
        sheet[f'F{index+11}'] = param.get('distribution')
        sheet[f'H{index+11}'] = param.get('params')

    permeability = stat_params.get('permeability', dict())
    sheet['E114'].value = permeability.get('distribution')
    sheet['G114'].value = permeability.get('params')

    init_data: dict = data.get('init_data', dict())
    sheet['K15'].value = init_data.get('relative_density')
    sheet['K16'].value = init_data.get('reservoir_temp')
    sheet['K17'].value = init_data.get('init_reservoir_pressure')
    sheet['K18'].value = init_data.get('temp_correction')
    sheet['K19'].value = init_data.get('init_overcompress_coef')
    sheet['F21'].value = init_data.get('num_of_vars')

    prod_profile_init_data: dict = data.get('prod_profile_init_data', dict())
    sheet['J117'].value = prod_profile_init_data.get('prod_rate')
    sheet['J118'].value = prod_profile_init_data.get('operations_ratio')
    sheet['J119'].value = prod_profile_init_data.get('reserve_ratio')
    sheet['J120'].value = prod_profile_init_data.get('machines_num')
    sheet['J121'].value = prod_profile_init_data.get('time_to_build')
    sheet['J123'].value = prod_profile_init_data.get('well_height')
    sheet['J124'].value = prod_profile_init_data.get('pipe_diameter')
    sheet['J125'].value = prod_profile_init_data.get('main_gas_pipeline_pressure')
    sheet['J127'].value = prod_profile_init_data.get('abandon_pressure')
    sheet['J130'].value = prod_profile_init_data.get('filtr_resistance_A')
    sheet['J131'].value = prod_profile_init_data.get('filtr_resistance_B')
    sheet['J132'].value = prod_profile_init_data.get('hydraulic_resistance')
    sheet['J133'].value = prod_profile_init_data.get('trail_length')
    sheet['J134'].value = prod_profile_init_data.get('input_cs_temp')

    risks_params: dict = data.get('risk_params', dict())
    sheet['H186'].value = risks_params.get('exploration_wells_amount')
    sheet['H187'].value = risks_params.get('distance_from_infra')

    risks_kriterias: dict = data.get('risks_kriterias', dict())
    seismic_exploration_work = risks_kriterias.get('seismic_exploration_work', dict())
    sheet['H190'].value = seismic_exploration_work.get('kriteria')
    sheet['J190'].value = seismic_exploration_work.get('value')
    sheet['K190'].value = seismic_exploration_work.get('weight')

    grid_density = risks_kriterias.get('grid_density', dict())
    sheet['H191'].value = grid_density.get('kriteria')
    sheet['J191'].value = grid_density.get('value')
    sheet['K191'].value = grid_density.get('weight')

    core_research = risks_kriterias.get('core_research', dict())
    sheet['H192'].value = core_research.get('kriteria')
    sheet['J192'].value = core_research.get('value')
    sheet['K192'].value = core_research.get('weight')

    c1_reserves = risks_kriterias.get('c1_reserves', dict())
    sheet['H193'].value = c1_reserves.get('kriteria')
    sheet['J193'].value = c1_reserves.get('value')
    sheet['K193'].value = c1_reserves.get('weight')

    hydrocarbon_properties = risks_kriterias.get('hydrocarbon_properties', dict())
    sheet['H194'].value = hydrocarbon_properties.get('kriteria')
    sheet['J194'].value = hydrocarbon_properties.get('value')
    sheet['K194'].value = hydrocarbon_properties.get('weight')

    study_coef = data.get('study_coef', None)
    sheet['F197'].value = study_coef

    profiles_report: pd.DataFrame = data['profiles_report']
    P90 = profiles_report.get('P90', dict())
    sheet['C25'].value = P90.get('geo_gas_reserves')

    P50 = profiles_report.get('P50', dict())
    sheet['C27'].value = P50.get('geo_gas_reserves')

    P10 = profiles_report.get('P10', dict())
    sheet['C29'].value = P10.get('geo_gas_reserves')

    for profile, col in zip(('P90', 'P50', 'P10'), ('F', 'H', 'J')):
        vars = profiles_report.get(profile, dict())
        sheet[f'{col}210'].value = vars.get('geo_gas_reserves')
        sheet[f'{col}211'].value = vars.get('area')
        sheet[f'{col}212'].value = vars.get('effective_thickness')
        sheet[f'{col}213'].value = vars.get('porosity_coef')
        sheet[f'{col}214'].value = vars.get('gas_saturation_coef')
        sheet[f'{col}215'].value = vars.get('relative_density')
        sheet[f'{col}216'].value = vars.get('reservoir_temp')
        sheet[f'{col}217'].value = vars.get('init_reservoir_pressure')
        sheet[f'{col}218'].value = vars.get('temp_correction')
        sheet[f'{col}219'].value = vars.get('init_overcompress_coef')
        sheet[f'{col}220'].value = vars.get('annual_production')
        sheet[f'{col}221'].value = vars.get('accumulated_production')
        sheet[f'{col}222'].value = vars.get('kig')
        sheet[f'{col}223'].value = vars.get('num_of_wells')
        sheet[f'{col}224'].value = vars.get('years')

    calcs_table: list[pd.DataFrame] = data.get('calcs_table', [])
    row_num = 228
    for df_table, name in zip(calcs_table, ["P10", "P50", "P90"]):
        sheet[f'A{row_num}'] = CellRichText(
            TextBlock(InlineFont(rFont="Times New Roman", b=False, sz=20), 'Расчёт показателей разработки - '),
            TextBlock(InlineFont(rFont="Times New Roman", b=True, sz=20), f'вариант {name}'),
        )
        sheet[f'A{row_num}'].alignment = Alignment(horizontal='center', vertical='center')
        sheet.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=len(df_table.columns))
        row_num += 1
        ins_row, _ = _insert_dataframe(df_table, sheet, row_num, index=False)
        row_num += ins_row + 2

    images = data.get('images', dict())
    for img_name, img_data in images.items():
        # Открываем изображение с помощью Pillow
        img = PILImage.open(io.BytesIO(img_data))

        # Сохраняем изображение во временный файл
        temp_img_path = TEMP_PATH+f"/temp_{field_name}_{img_name}.png"
        img.save(temp_img_path)

        # Вставляем изображение в Excel
        img_excel = Image(temp_img_path)
        if img_name == "hist":
            sheet.add_image(img_excel, 'A54')
        elif img_name == "tornado":
            sheet.add_image(img_excel, 'A82')
        elif img_name == "profile":
            sheet.add_image(img_excel, 'A157')


def insert_comparison(data: dict, sheet: wsxl.Worksheet):
    if (data.get('comparison_values') is None
            or data.get('comparison_images') is None
            or data.get('comparison_values', pd.DataFrame()).empty):
        return

    # pdb.set_trace()
    comparison_values = data['comparison_values']
    indexes = [
        VARNAMES_ANALYSIS['geo_gas_reserves'],
        VARNAMES_ANALYSIS['study_coef'],
        VARNAMES_ANALYSIS['uncertainty_coef'],
        VARNAMES_ANALYSIS['annual_production'],
        VARNAMES_ANALYSIS['distance_from_infra'],
        VARNAMES_ANALYSIS['accumulated_production'],
    ]
    comparison_values = comparison_values.reindex(indexes)

    rows = dataframe_to_rows(comparison_values, header=True, index=False)

    for row_id, row in enumerate(rows, 1):
        for col_id, value in enumerate(row, 1):
            sheet.cell(row=row_id, column=col_id+1, value = value).border = Border(
                right=Side(border_style='thin', color='FF000000'),
                left=Side(border_style='thin', color='FF000000'),
                top=Side(border_style='thin', color='FF000000'),
                bottom=Side(border_style='thin', color='FF000000'),
            )
            sheet[f'{chr(ord('A')+col_id)}{row_id}'].font = Font(name='Times New Roman', size=14)
            sheet[f'{chr(ord('A')+col_id)}{row_id}'].alignment = Alignment(horizontal='center',
                                                                           vertical='center')


    for column_cells in sheet.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = length*EXCEL_CELL_LENGTH_COEF

    images = data.get('comparison_images', dict())
    for img_name, img_data in images.items():
        img = PILImage.open(io.BytesIO(img_data))
        temp_img_path = TEMP_PATH+f"/temp_{img_name}.png"
        img.save(temp_img_path)

        img_excel = Image(temp_img_path)

        if img_name == 'study_coef_chart':
            sheet.add_image(img_excel, 'A9')
        if img_name == 'uncertainty_coef_chart':
            sheet.add_image(img_excel, 'A35')
        if img_name == 'annual_production_chart':
            sheet.add_image(img_excel, 'A61')
        if img_name == 'distance_from_infra_chart':
            sheet.add_image(img_excel, 'A87')

def _insert_dataframe(data: pd.DataFrame,
                      sheet: wsxl.Worksheet,
                      row_start: int = 1,
                      col_start: int = 1,
                      header: bool = True,
                      index: bool = True,
                      ) -> tuple[int, int]:
    rows = dataframe_to_rows(data, header=header, index=index)
    row_correction = 0
    for row_id, row in enumerate(rows, 0):
        if index and row_id == 1:
            row_correction = -1
            continue
        row_id += row_start + row_correction
        for col_id, value in enumerate(row, col_start):
            sheet.cell(row=row_id, column=col_id, value=value).border = Border(
                right=Side(border_style='thin', color='FF000000'),
                left=Side(border_style='thin', color='FF000000'),
                top=Side(border_style='thin', color='FF000000'),
                bottom=Side(border_style='thin', color='FF000000'),
            )
            sheet[f'{chr(ord('A') + col_id - 1)}{row_id}'].font = Font(name='Times New Roman', size=14)
            sheet[f'{chr(ord('A') + col_id - 1)}{row_id}'].alignment = Alignment(horizontal='center', vertical='center')
    return len(data)+1, len(data.columns)+1


def insert_fields_comparison(data: dict, sheet: wsxl.Worksheet):
    if (data.get('summ_tables') is None
            or data.get('selected_fields_tables') is None
            or data.get('chart') is None):
        return

    # pdb.set_trace()
    summ_tables: dict[str, tuple[pd.DataFrame, list[str]]] = data['summ_tables']
    selected_fields_tables: dict[str, pd.DataFrame] = data['selected_fields_tables']
    indexes = [
        VARNAMES_ANALYSIS['geo_gas_reserves'],
        VARNAMES_ANALYSIS['study_coef'],
        VARNAMES_ANALYSIS['uncertainty_coef'],
        VARNAMES_ANALYSIS['annual_production'],
        VARNAMES_ANALYSIS['accumulated_production'],
    ]

    row_index = 1
    col_index = 1

    for group, (df, selected) in summ_tables.items():
        summ_table = df.reindex(indexes)

        sheet.cell(row=row_index, column=col_index).value = f"Итоговая таблица {group} ({', '.join(field for field in selected)})"
        sheet.cell(row=row_index, column=col_index).font = Font(name='Times New Roman', size=14)
        sheet.cell(row=row_index, column=col_index).alignment = Alignment(horizontal='center', vertical='center')
        sheet.merge_cells(start_row=row_index, start_column=col_index, end_row=row_index, end_column=len(summ_table.columns)+1)
        row_index += 2
        col_index = 1

        ins_rows, ins_cols = _insert_dataframe(summ_table, sheet, row_start=row_index, col_start=col_index)
        row_index += ins_rows


        row_index += 3

    for field in selected_fields_tables:
        df: pd.DataFrame = selected_fields_tables[field].reindex(indexes)
        sheet[f'A{row_index+1}'].value = field
        sheet[f'A{row_index+1}'].font = Font(name='Times New Roman', size=14)
        sheet[f'A{row_index+1}'].alignment = Alignment(horizontal='center', vertical='center')
        row_index += 3
        ins_rows, ins_cols = _insert_dataframe(df, sheet, row_start=row_index)
        row_index += ins_rows
        col_index = ins_cols + 2

    length = max(len(ind) for ind in indexes)
    sheet.column_dimensions['A'].width = length*EXCEL_CELL_LENGTH_COEF


    chart = data.get('chart', bytes())
    img = PILImage.open(io.BytesIO(chart))
    temp_img_path = TEMP_PATH + f"/temp_chart_fields_comparison.png"
    img.save(temp_img_path)

    img_excel = Image(temp_img_path)

    sheet.add_image(img_excel, f'{chr(ord('A')+col_index+1)}1')




def create_report(excel_data: dict, template_path: str, output_path: str):
    data = excel_data['fields']
    prod_sites = data.keys()

    template_wb = xl.load_workbook(template_path)
    output_wb = xl.Workbook()

    # Удаляем лист по умолчанию, который создается при создании нового рабочего файла
    if 'Sheet' in output_wb.sheetnames:
        del output_wb['Sheet']

    for field in prod_sites:
        copy_sheet(template_wb, 'Sheet1', output_wb, field)
        insert_to_sheet(field, data[field], output_wb[field])

    copy_sheet(template_wb, 'ComparisonAnalysis', output_wb, 'Сравнительный Анализ')
    insert_comparison(excel_data['comparison'], output_wb['Сравнительный Анализ'])

    output_wb.create_sheet('Сравнение месторождений')
    insert_fields_comparison(excel_data['fields_comparison'], output_wb['Сравнение месторождений'])

    output_wb.save(output_path)

    print(f'Файл {output_path} успешно создан.')


