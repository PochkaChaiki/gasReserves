import openpyxl as xl
# from openpyxl.styles import *
import openpyxl.worksheet.worksheet as wsxl
import pandas as pd
from PIL import Image as PILImage
from openpyxl.drawing.image import Image
import io
import os

from openpyxl.utils.dataframe import dataframe_to_rows


def copy_sheet(source_wb: xl.Workbook, source_sheet_name: wsxl.Worksheet, target_wb: xl.Workbook, target_sheet_name: wsxl.Worksheet):
    source_sheet = source_wb[source_sheet_name]
    target_sheet = target_wb.create_sheet(title=target_sheet_name)

    for row in source_sheet.iter_rows(values_only=False):
        for cell in row:
            new_cell = target_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = xl.styles.Font(**cell.font.__dict__)
                new_cell.border = xl.styles.Border(**cell.border.__dict__)
                new_cell.fill = xl.styles.PatternFill(**cell.fill.__dict__)
                # new_cell.number_format = xl.styles.NumberFormat(**cell.number_format.__dict__)
                new_cell.protection = xl.styles.Protection(**cell.protection.__dict__)
                new_cell.alignment = xl.styles.Alignment(**cell.alignment.__dict__)

    for merged_cell in source_sheet.merged_cells:
        target_sheet.merge_cells(range_string=merged_cell.coord)


'''
---------------------------------------------------------------------------------------------
Data dict: 
data = {
    "placement1": {
        "stat_params": {
            "var": {
                "distribution": str,
                "params": str,
            },    
        },
        "init_data": dict,
        "prod_profile_init_data": dict,
        "risks_params": dict,
        "risks_kriterias": {
            "seismic...": {
                "kriteria":,
                "value":,
                "weight":,
            },
            ...
        },
        "study_coef": float,
        "profiles_report": pd.DataFrame,
        "images": {"hist": bytes, "tornado": bytes, "profile": bytes},
    }, 
    "placement2": {
        ...
    },
    "placement3": {
        ...
    },
    ...
}
---------------------------------------------------------------------------------------------
'''
def insert_to_sheet(data: dict, sheet: wsxl.Worksheet):
    vars = ['area', 'effective_thickness', 'porosity_coef', 'gas_saturation_coef']
    stat_params: dict = data['stat_params']
    for index in range(0, 4, 1):
        sheet[f'F{index+11}'] = stat_params[vars[index]]['distribution']
        sheet[f'H{index+11}'] = stat_params[vars[index]]['params']
    
    sheet['E114'].value = stat_params['permeability']['distribution']
    sheet['G114'].value = stat_params['permeability']['params']

    init_data: dict = data['init_data']
    sheet['K15'].value = init_data['relative_density']
    sheet['K16'].value = init_data['reservoir_temp']
    sheet['K17'].value = init_data['init_reservoir_pressure']
    sheet['K18'].value = init_data['temp_correction']
    sheet['K19'].value = init_data['init_overcompress_coef']
    sheet['F21'].value = init_data['num_of_vars']

    prod_profile_init_data: dict = data['prod_profile_init_data']
    sheet['J117'].value = prod_profile_init_data['prod_rate']
    sheet['J118'].value = prod_profile_init_data['operations_ratio']
    sheet['J119'].value = prod_profile_init_data['reserve_ratio']
    sheet['J120'].value = prod_profile_init_data['machines_num']
    sheet['J121'].value = prod_profile_init_data['time_to_build']
    sheet['J123'].value = prod_profile_init_data['well_height']
    sheet['J124'].value = prod_profile_init_data['pipe_diameter']
    sheet['J125'].value = prod_profile_init_data['main_gas_pipeline_pressure']
    sheet['J127'].value = prod_profile_init_data['abandon_pressure']
    sheet['J130'].value = prod_profile_init_data['filtr_resistance_A']
    sheet['J131'].value = prod_profile_init_data['filtr_resistance_B']
    sheet['J132'].value = prod_profile_init_data['hydraulic_resistance']
    sheet['J133'].value = prod_profile_init_data['trail_length']
    sheet['J134'].value = prod_profile_init_data['input_cs_temp']

    risks_params: dict = data['risk_params']
    sheet['H186'].value = risks_params['exploration_wells_amount']
    sheet['H187'].value = risks_params['distance_from_infra']

    risks_kriterias: dict = data['risks_kriterias']
    sheet['H190'].value = risks_kriterias['seismic_exploration_work']['kriteria']
    sheet['J190'].value = risks_kriterias['seismic_exploration_work']['value']
    sheet['K190'].value = risks_kriterias['seismic_exploration_work']['weight']

    sheet['H191'].value = risks_kriterias['grid_density']['kriteria']
    sheet['J191'].value = risks_kriterias['grid_density']['value']
    sheet['K191'].value = risks_kriterias['grid_density']['weight']

    sheet['H192'].value = risks_kriterias['core_research']['kriteria']
    sheet['J192'].value = risks_kriterias['core_research']['value']
    sheet['K192'].value = risks_kriterias['core_research']['weight']

    sheet['H193'].value = risks_kriterias['c1_reserves']['kriteria']
    sheet['J193'].value = risks_kriterias['c1_reserves']['value']
    sheet['K193'].value = risks_kriterias['c1_reserves']['weight']

    sheet['H194'].value = risks_kriterias['hydrocarbon_properties']['kriteria']
    sheet['J194'].value = risks_kriterias['hydrocarbon_properties']['value']
    sheet['K194'].value = risks_kriterias['hydrocarbon_properties']['weight']

    study_coef: float = data['study_coef']
    sheet['F197'].value = study_coef

    profiles_report: pd.DataFrame = data['profiles_report']
    sheet['C25'].value = profiles_report['P90']['geo_gas_reserves']
    sheet['C27'].value = profiles_report['P50']['geo_gas_reserves']
    sheet['C29'].value = profiles_report['P10']['geo_gas_reserves']
    for profile, col in zip(('P90', 'P50', 'P10'), ('F', 'H', 'J')):
        vars = profiles_report[profile]
        sheet[f'{col}210'].value = vars['geo_gas_reserves']
        sheet[f'{col}211'].value = vars['effective_thickness']
        sheet[f'{col}212'].value = vars['porosity_coef']
        sheet[f'{col}213'].value = vars['gas_saturation_coef']
        sheet[f'{col}214'].value = vars['relative_density']
        sheet[f'{col}215'].value = vars['reservoir_temp']
        sheet[f'{col}216'].value = vars['init_reservoir_pressure']
        sheet[f'{col}217'].value = vars['temp_correction']
        sheet[f'{col}218'].value = vars['init_overcompress_coef']
        sheet[f'{col}219'].value = vars['annual_production']
        sheet[f'{col}220'].value = vars['accumulated_production']
        sheet[f'{col}221'].value = vars['kig']
        sheet[f'{col}222'].value = vars['num_of_wells']
        sheet[f'{col}223'].value = vars['years']
    
    images = data['images']
    for img_name, img_data in images.items():
        # Открываем изображение с помощью Pillow
        img = PILImage.open(io.BytesIO(img_data))

        # Сохраняем изображение во временный файл
        temp_img_path = f"./temp/temp_{img_name}.png"
        img.save(temp_img_path)

        # Вставляем изображение в Excel
        img_excel = Image(temp_img_path)
        if img_name == "hist":
            sheet.add_image(img_excel, 'A54')
        elif img_name == "tornado":
            sheet.add_image(img_excel, 'A82')
        elif img_name == "profile":
            sheet.add_image(img_excel, 'A157')
        # os.remove(temp_img_path)


def insert_comparison(data: dict, sheet: wsxl.Worksheet):
    rows = dataframe_to_rows(data['comparison_values'], header=True, index=False)

    for row_id, row in enumerate(rows, 1):
        for col_id, value in enumerate(row, 1):
            sheet.cell(row=row_id, column=col_id+4, value = value)

    images = data['comparison_images']
    for img_name, img_data in images.items():
        img = PILImage.open(io.BytesIO(img_data))
        temp_img_path = f"./temp/temp_{img_name}.png"
        img.save(temp_img_path)

        img_excel = Image(temp_img_path)

        if img_name == 'study_coef_chart':
            sheet.add_image(img_excel, 'A9')
        if img_name == 'uncertainty_coef_chart':
            sheet.add_image(img_excel, 'M9')
        if img_name == 'annual_production_chart':
            sheet.add_image(img_excel, 'A35')
        if img_name == 'distance_from_infra_chart':
            sheet.add_image(img_excel, 'M35')


def create_report(excel_data: dict, template_path: str, output_path: str):
    data = excel_data['fields']
    prod_sites = data.keys()

    template_wb = xl.load_workbook(template_path)
    output_wb = xl.Workbook()

    # Удаляем лист по умолчанию, который создается при создании нового рабочего файла
    if 'Sheet' in output_wb.sheetnames:
        del output_wb['Sheet']

    for placement in prod_sites:
        copy_sheet(template_wb, 'Sheet1', output_wb, placement)
        insert_to_sheet(data[placement], output_wb[placement])

    copy_sheet(template_wb, 'ComparisonAnalysis', output_wb, 'Сравнительный Анализ')
    insert_comparison(excel_data['comparison'], output_wb['Сравнительный Анализ'])

    output_wb.save(output_path)

    # print(f'Файл {output_path} успешно создан.')


